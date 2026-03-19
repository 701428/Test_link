import base64
import io
import json
from urllib.parse import urlparse, parse_qs, urlencode, quote, urlunparse

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from openpyxl.styles import Font

app = FastAPI(title="SharePoint Test Report Linker")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Results"],
)


@app.get("/health")
@app.get("/api/health")
async def health():
    return {"status": "ok"}


@app.post("/process")
@app.post("/api/process")
async def process_excel(
    file: UploadFile = File(...),
    base_url: str = Form(...),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    # Build a function to construct per-sheet URLs correctly
    parsed = urlparse(base_url)
    qs = parse_qs(parsed.query, keep_blank_values=True)

    def sheet_url(sheet_name: str) -> str:
        # SharePoint OneDrive URLs use ?id=<encoded-path>&ga=1
        # parse_qs decodes the id value, so we re-encode it manually
        # to avoid double-encoding the existing slashes.
        if "id" in qs:
            base_path = qs["id"][0].rstrip("/")
            full_path = base_path + "/" + sheet_name
            # quote the full path: keep / as %2F (safe='') to match SharePoint format
            encoded_path = quote(full_path, safe="")
            # Rebuild remaining query params (e.g. ga=1) excluding id
            other = {k: v for k, v in qs.items() if k != "id"}
            other_str = ("&" + urlencode(other, doseq=True)) if other else ""
            new_query = "id=" + encoded_path + other_str
            return urlunparse(parsed._replace(query=new_query))
        # Fallback: plain URL, just append sheet name
        return base_url.rstrip("/") + "/" + sheet_name

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        linked = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "test report path":
                    target_row = cell.row
                    target_col = cell.column + 1
                    folder_url = sheet_url(sheet_name)

                    # Unmerge any merged region that covers the target cell
                    merged_to_remove = [
                        mr for mr in ws.merged_cells.ranges
                        if target_row >= mr.min_row and target_row <= mr.max_row
                        and target_col >= mr.min_col and target_col <= mr.max_col
                    ]
                    for mr in merged_to_remove:
                        ws.unmerge_cells(str(mr))

                    target = ws.cell(row=target_row, column=target_col)
                    target.value = sheet_name
                    target.hyperlink = folder_url
                    target.font = Font(color="0563C1", underline="single")

                    print(f"[linked] {sheet_name} -> {folder_url}")
                    results.append({"sheet": sheet_name, "url": folder_url, "status": "linked"})
                    linked = True
                    break
            if linked:
                break

        if not linked:
            results.append({"sheet": sheet_name, "url": None, "status": "Test Report path label not found"})

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = file.filename.replace(".xlsx", "_linked.xlsx")
    file_b64 = base64.b64encode(output.read()).decode("utf-8")

    return JSONResponse({
        "filename": filename,
        "file": file_b64,
        "results": results,
    })
